"""
xlsb_to_csv.py
==============
Converte arquivos .xlsb e .xlsx para CSVs usando um pipeline de 3 camadas:

  Camada 1 — pandas (local, sem API)
      Detecta tabela limpa ou com header deslocado automaticamente.
      Resolve a maioria dos casos comuns sem custo de API.

  Camada 2 — heurística local (sem API)
      Varre linhas buscando a com mais valores não-nulos como header.
      Detecta múltiplas tabelas separadas por blocos de linhas vazias.

  Camada 3 — Claude (fallback explícito)
      Chamado apenas se as duas camadas anteriores falharem.
      Falha do Claude levanta exceção — tolerância zero a silêncio.

Dependências:
    pip install anthropic pandas pyxlsb openpyxl

Uso como módulo:
    from xlsb_to_csv import convert_file
    reports = convert_file("relatorio.xlsb", output_dir="saida/")

Uso como CLI:
    python xlsb_to_csv.py relatorio.xlsb --output-dir saida/
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import os
import re
import sys
import time
import traceback
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Any
from utils.constants import API_KEY
import anthropic
import openpyxl
import pandas as pd
import pyxlsb
from utils.reporter_base import success, failure, Report

logger = logging.getLogger(__name__)

# ── Constantes ────────────────────────────────────────────────────────────────

MODEL            = "claude-haiku-4-5-20251001"
MAX_TOKENS       = 8192
SMALL_SHEET_ROWS = 300   # abas com até N linhas → Claude envia tudo (se chegar lá)
PREVIEW_ROWS     = 60    # linhas de preview para Claude em abas grandes
MAX_HEADER_SCAN  = 20    # máx de linhas para procurar header na heurística
MIN_FILL_RATIO   = 0.5   # proporção mínima de células preenchidas para linha ser header
MAX_UNNAMED_RATIO = 0.2  # tolerância de colunas "Unnamed:*" no pandas
MIN_DATA_ROWS    = 2     # mínimo de linhas de dados para bloco ser considerado tabela
HEADER_ROW_TRESHOLD = 0.8

# ── Tipos de resultado ────────────────────────────────────────────────────────

class Layer(str, Enum):
    PANDAS    = "pandas"
    HEURISTIC = "heuristic"
    CLAUDE    = "claude"


@dataclass
class TableResult:
    """Resultado da extração de uma tabela lógica dentro de uma aba."""
    csv_path:      Path
    sheet_name:    str
    table_name:    str
    columns:       list[str]
    row_count:     int
    layer_used:    Layer
    elapsed_s:     float
    warnings:      list[str] = field(default_factory=list)


@dataclass
class SheetError:
    """Erro fatal ao processar uma aba — sem silêncio."""
    sheet_name: str
    reason:     str
    layer:      Layer
    traceback:  str = ""


# ── Leitura bruta ─────────────────────────────────────────────────────────────

def _sheet_names(wb_path: Path) -> list[str]:
    ext = wb_path.suffix.lower()
    if ext == ".xlsb":
        with pyxlsb.open_workbook(str(wb_path)) as wb:
            return list(wb.sheets)
    if ext == ".xlsx":
        wb = openpyxl.load_workbook(str(wb_path), read_only=True)
        names = list(wb.sheetnames)
        wb.close()
        return names
    raise ValueError(f"Formato não suportado: {ext}. Use .xlsb ou .xlsx")


def _read_raw(wb_path: Path, sheet_name: str) -> list[list[Any]]:
    """Retorna lista de listas com os valores brutos da aba."""
    ext = wb_path.suffix.lower()
    if ext == ".xlsb":
        rows: list[list[Any]] = []
        with pyxlsb.open_workbook(str(wb_path)) as wb:
            with wb.get_sheet(sheet_name) as ws:
                for row in ws.rows():
                    rows.append([c.v for c in row])
        return rows
    if ext == ".xlsx":
        wb = openpyxl.load_workbook(str(wb_path), read_only=True, data_only=True)
        ws = wb[sheet_name]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
        return rows
    raise ValueError(f"Formato não suportado: {ext}")


def _trim_trailing_none(row: list[Any]) -> list[Any]:
    last = -1
    for i, v in enumerate(row):
        if v is not None:
            last = i
    return row[: last + 1] if last >= 0 else []


def _fill_ratio(row: list[Any]) -> float:
    if not row:
        return 0.0
    return sum(1 for v in row if v is not None) / len(row)


def _is_blank_row(row: list[Any]) -> bool:
    return all(v is None or str(v).strip() == "" for v in row)


# ── Camada 1: pandas ──────────────────────────────────────────────────────────

def _unnamed_ratio(df: pd.DataFrame) -> float:
    unnamed = sum(1 for c in df.columns if str(c).startswith("Unnamed:"))
    return unnamed / max(len(df.columns), 1)


def _detect_blocks(raw_rows: list[list[Any]]) -> list[tuple[int, int]]:
    """
    Divide linhas brutas em blocos contínuos não-vazios.
    Retorna lista de (start_idx, end_idx) inclusive.
    """
    blocks: list[tuple[int, int]] = []
    in_block = False
    start = 0
    for i, row in enumerate(raw_rows):
        if not _is_blank_row(row):
            if not in_block:
                start = i
                in_block = True
        else:
            if in_block:
                blocks.append((start, i - 1))
                in_block = False
    if in_block:
        blocks.append((start, len(raw_rows) - 1))
    return blocks


def _block_to_dataframe(
    raw_rows: list[list[Any]],
    start: int,
    end: int,
    warnings: list[str],
    block_idx: int,
) -> "pd.DataFrame | None":
    block = raw_rows[start: end + 1]
    if not block:
        return None

    best_ratio = -1.0
    best_idx = 0
    for i, row in enumerate(block[:MAX_HEADER_SCAN]):
        trimmed = _trim_trailing_none(row)
        r = _fill_ratio(trimmed)
        if r > HEADER_ROW_TRESHOLD:
            best_ratio = r
            best_idx = i
            break
        if r > best_ratio:
            best_ratio = r
            best_idx = i
    
    if best_ratio < MIN_FILL_RATIO:
        return None

    if best_idx > 0 or block_idx > 0:
        warnings.append(
            f"Bloco {block_idx + 1}: header na linha relativa {best_idx} "
            f"(absoluta {start + best_idx})"
        )

    header_raw = _trim_trailing_none(block[best_idx])
    columns = [
        str(v).strip() if v is not None else f"CAMPO_{i}"
        for i, v in enumerate(header_raw)
    ]
    n_cols = len(columns)

    data_rows = []
    for row in block[best_idx + 1:]:
        if _is_blank_row(row):
            continue
        padded = list(row) + [None] * max(0, n_cols - len(row))
        data_rows.append(padded[:n_cols])

    if not data_rows:
        return None

    df = pd.DataFrame(data_rows, columns=columns)
    df = df.astype(str).replace("None", pd.NA).dropna(how="all")
    return df if not df.empty else None


def _try_pandas(
    wb_path: Path,
    sheet_name: str,
    raw_rows: list[list[Any]] | None = None,
) -> "tuple[list[pd.DataFrame], list[str]] | None":
    """
    Tenta extrair tabelas via raw rows + pandas.

    1. Detecta blocos contínuos separados por linhas vazias.
    2. Para cada bloco, varre as primeiras linhas em busca do melhor header.
    3. Valida unnamed_ratio; retorna None se nenhum bloco for aceitável.
    """
    warnings: list[str] = []

    if raw_rows is None:
        raw_rows = _read_raw(wb_path, sheet_name)

    if not raw_rows:
        return None

    blocks = _detect_blocks(raw_rows)
    if not blocks:
        return None
    tables: list[pd.DataFrame] = []
    for i, (start, end) in enumerate(blocks):
        df = _block_to_dataframe(raw_rows, start, end, warnings, i)
        if df is not None:
            if len(df) < MIN_DATA_ROWS:
                warnings.append(f"Bloco {i + 1} ignorado: apenas {len(df)} linha(s) de dados (mínimo={MIN_DATA_ROWS})")
            elif _unnamed_ratio(df) > MAX_UNNAMED_RATIO:
                warnings.append(
                    f"Bloco {i + 1} ignorado: unnamed_ratio={_unnamed_ratio(df):.0%}"
                )
            else:
                tables.append(df)

    if not tables:
        return None

    if len(blocks) > 1:
        warnings.append(f"{len(tables)} tabela(s) detectada(s) na aba")

    return tables, warnings


# ── Camada 2: heurística local ────────────────────────────────────────────────

def _try_heuristic(
    raw_rows: list[list[Any]],
) -> tuple[list[dict], list[str]] | None:
    """
    Detecta tabelas sem API.

    Algoritmo:
      1. Varre as primeiras MAX_HEADER_SCAN linhas buscando a com mais
         valores não-nulos → candidata a header.
      2. Dados começam logo após; gaps de linhas vazias delimitam tabelas.
      3. Se header_row tiver fill_ratio < MIN_FILL_RATIO → retorna None.

    Retorna lista de dicts {"header": [...], "rows": [[...]]} + warnings,
    ou None se não conseguiu detectar.
    """
    warnings: list[str] = []

    if not raw_rows:
        return None

    # Encontra melhor header candidate
    best_row_idx = 0
    best_fill = 0.0
    for i, row in enumerate(raw_rows[:MAX_HEADER_SCAN]):
        trimmed = _trim_trailing_none(row)
        ratio = _fill_ratio(trimmed)
        if ratio > best_fill:
            best_fill = ratio
            best_row_idx = i

    if best_fill < MIN_FILL_RATIO:
        warnings.append(
            f"Melhor candidata a header (linha {best_row_idx}) tem fill_ratio={best_fill:.2f} < {MIN_FILL_RATIO}"
        )
        return None

    header = [
        str(v).strip() if v is not None else f"CAMPO_{i}"
        for i, v in enumerate(_trim_trailing_none(raw_rows[best_row_idx]))
    ]
    data_start = best_row_idx + 1

    if best_row_idx > 0:
        warnings.append(f"Header detectado na linha {best_row_idx} pela heurística local")

    # Divide dados por gaps de linhas vazias
    tables: list[dict] = []
    current_rows: list[list[Any]] = []
    current_header = header

    for i in range(data_start, len(raw_rows)):
        row = raw_rows[i]
        if _is_blank_row(row):
            if current_rows:
                tables.append({"header": current_header, "rows": current_rows})
                current_rows = []
                # Tenta detectar novo header após gap
                current_header = None  # type: ignore[assignment]
        else:
            if current_header is None:
                # Linha após gap vira novo header
                trimmed = _trim_trailing_none(row)
                if _fill_ratio(trimmed) >= MIN_FILL_RATIO:
                    current_header = [
                        str(v).strip() if v is not None else f"CAMPO_{j}"
                        for j, v in enumerate(trimmed)
                    ]
                    warnings.append(f"Nova tabela detectada a partir da linha {i}")
                    continue
                else:
                    current_header = header  # fallback para header original
            trimmed = row[: len(current_header)]
            if any(v is not None for v in trimmed):
                current_rows.append(trimmed)

    if current_rows:
        tables.append({"header": current_header or header, "rows": current_rows})

    if not tables:
        return None

    return tables, warnings


# ── Camada 3: Claude ──────────────────────────────────────────────────────────

def _rows_to_text(rows: list[list[Any]], max_rows: int | None = None) -> str:
    subset = rows[:max_rows] if max_rows else rows
    lines = []
    for i, row in enumerate(subset):
        trimmed = _trim_trailing_none(row)
        lines.append(f"row {i}: {json.dumps(trimmed, ensure_ascii=False, default=str)}")
    return "\n".join(lines)


def _call_claude(client: anthropic.Anthropic, prompt: str, max_tokens=MAX_TOKENS) -> str:
    msg = client.messages.create(
        model=MODEL,
        max_tokens=max_tokens,
        messages=[{"role": "user", "content": prompt}],
    )
    return msg.content[0].text


def _extract_json(text: str) -> Any:
    match = re.search(r"```(?:json)?\s*([\s\S]+?)```", text)
    if match:
        return json.loads(match.group(1))
    match = re.search(r"(\{[\s\S]*\}|\[[\s\S]*\])", text)
    if match:
        return json.loads(match.group(1))
    raise ValueError(f"Nenhum JSON encontrado na resposta do Claude:\n{text[:500]}")


def _try_claude(
    client: anthropic.Anthropic,
    sheet_name: str,
    raw_rows: list[list[Any]],
) -> list[dict]:
    """
    Fallback final via Claude. Levanta exceção se falhar — sem silêncio.

    Retorna lista de dicts {"header": [...], "rows": [[...]]}
    """
    is_large = len(raw_rows) > SMALL_SHEET_ROWS

    if is_large:
        rows_text = _rows_to_text(raw_rows, max_rows=PREVIEW_ROWS)
        prompt = f"""Você receberá um preview de uma aba de planilha chamada "{sheet_name}".
Identifique TODAS as tabelas tabulares. Para cada uma informe:
  - header_row: índice da linha com os cabeçalhos
  - data_start_row: índice da primeira linha de dados
  - data_end_row: índice da última linha de dados (null = até o fim)

Retorne SOMENTE JSON válido, sem texto adicional:
{{
  "tables": [
    {{
      "name": "<nome_sem_espacos>",
      "header_row": <int>,
      "data_start_row": <int>,
      "data_end_row": <int|null>
    }}
  ]
}}

Preview:
{rows_text}
"""
        response = _call_claude(client, prompt)
        spec = _extract_json(response)
        tables_spec = spec.get("tables", [])

        tables: list[dict] = []
        for t in tables_spec:
            h = t.get("header_row")
            ds = t.get("data_start_row")
            de = t.get("data_end_row")
            if h is None or ds is None:
                raise ValueError(
                    f"Claude retornou spec inválida para aba '{sheet_name}': {t}"
                )
            header = [
                str(v).strip() if v is not None else f"CAMPO_{i}"
                for i, v in enumerate(_trim_trailing_none(raw_rows[h]))
            ]
            data: list[list[Any]] = []
            for i in range(ds, len(raw_rows)):
                if de is not None and i > de:
                    break
                row = raw_rows[i][: len(header)]
                if any(v is not None for v in row):
                    data.append(row)
            tables.append({"header": header, "rows": data})
        return tables

    else:
        rows_text = _rows_to_text(raw_rows)
        prompt = f"""Você receberá o conteúdo bruto de uma aba chamada "{sheet_name}".
Extraia TODAS as tabelas lógicas. Ignore títulos, notas de rodapé e linhas descritivas.

Retorne SOMENTE JSON válido, sem texto adicional:
{{
  "tables": [
    {{
      "name": "<nome_sem_espacos>",
      "columns": ["<col1>", ...],
      "rows": [[val, ...], ...]
    }}
  ]
}}

Conteúdo:
{rows_text}
"""
        response = _call_claude(client, prompt)
        spec = _extract_json(response)
        tables_spec = spec.get("tables", [])

        tables = []
        for t in tables_spec:
            cols = t.get("columns", [])
            rows = t.get("rows", [])
            if not cols:
                raise ValueError(
                    f"Claude retornou tabela sem colunas para aba '{sheet_name}': {t}"
                )
            tables.append({"header": cols, "rows": rows})
        return tables


# ── Extração de data de referência ───────────────────────────────────────────

def _extract_ref_date(path: str, sheet_name:str, client: anthropic.Anthropic | None = None) -> tuple[str, str]:
    """
    Extrai ano e mês do nome do arquivo.
    Tenta regex local primeiro; usa Claude como fallback.

    Retorna (ano, mes) como strings, ex: ("2022", "01") ou ("0000", "00").
    """
    

    prompt = (
        f'A partir do nome de arquivo "{path}" e do sheet_name {sheet_name}, retorne mês e ano de referência '
        f'no formato YYYY_MM. Sem comentários. Se não encontrar, use 00 no campo faltante. '
        f'Se for range de anos, use YYYY-YYYY.'
        f'Se for range de meses, use YYYY_01-12.'
        f'Não retorne aspas, listras, virgulas nem nada além do essencial'
        f'Exemplo dos ÚNICOS formatos de retornos válidos: 1. 2025_02\n 2. 2022-2023\n 3. 2012_01-03'
    )
    try:
        result = _call_claude(client, prompt, 10).strip()
        parts = result.split("_")
        return parts[0], parts[1] if len(parts) > 1 else "00"
    except Exception as e:
        raise RuntimeError(
            f"Claude falhou ao extrair data de referência de '{path}': {e}"
        ) from e


# ── Escrita de CSV ─────────────────────────────────────────────────────────────

def _sanitize(name: str) -> str:
    s = re.sub(r"[^\w\-]", "_", name)
    s = re.sub(r"_+", "_", s).strip("_")
    return s or "tabela"


def _write_csv(
    output_dir: Path,
    ano: str,
    mes: str,
    sheet_slug: str,
    table_name: str,
    columns: list[str],
    data: list[list[Any]],
) -> Path:
    dest = output_dir / ano / mes
    dest.mkdir(parents=True, exist_ok=True)
    out_path = dest / f"{sheet_slug}__{table_name}.csv"
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(columns)
        writer.writerows(data)
    return out_path


# ── Orquestrador principal ────────────────────────────────────────────────────

def _process_sheet(
    wb_path: Path,
    sheet_name: str,
    output_dir: Path,
    ano: str,
    mes: str,
    client: anthropic.Anthropic | None,
) -> list[TableResult | SheetError]:
    """Pipeline de 3 camadas para uma única aba."""
    sheet_slug = _sanitize(sheet_name)
    results: list[Report] = []

    # Lê raw_rows uma única vez — compartilhado entre todas as camadas
    try:
        raw_rows = _read_raw(wb_path, sheet_name)
    except Exception as e:
        return [failure(reason=f"Erro ao ler aba: {e}", solution="Verifique se a aba realmente existe e se não está corrompida", file_path=wb_path),]
    

    # ── Camada 1: pandas ──────────────────────────────────────────────────────
    t0 = time.perf_counter()
    try:
        pandas_result = _try_pandas(wb_path, sheet_name, raw_rows=raw_rows)
    except Exception as e:
        pandas_result = None
        

    if pandas_result is not None:
        tables_df, warnings = pandas_result
        for i, df in enumerate(tables_df):
            t_name = f"tabela_{i + 1}" if len(tables_df) > 1 else "tabela"
            columns = list(df.columns)
            data = df.values.tolist()
            out = _write_csv(output_dir, ano, mes, sheet_slug, t_name, columns, data)
            results.append(
                success(
                    file_path=out,
                    find_cols=columns,
                    execution_time_seconds=round(time.perf_counter() - t0, 4),
                    warning=warnings
                )
            )
        return results

    # ── Camada 2: heurística local ────────────────────────────────────────────
    t0 = time.perf_counter()
    try:
        heuristic_result = _try_heuristic(raw_rows)
    except Exception as e:
        heuristic_result = None

    if heuristic_result is not None:
        tables_h, warnings = heuristic_result
        for i, t in enumerate(tables_h):
            t_name = f"tabela_{i + 1}" if len(tables_h) > 1 else "tabela"
            out = _write_csv(
                output_dir, ano, mes, sheet_slug, t_name,
                t["header"], t["rows"],
            )
            results.append(
                success(
                    file_path=out,
                    find_cols=columns,
                    execution_time_seconds=round(time.perf_counter() - t0, 4),
                    warning="Foi utilizado o método heurístico, que pode apresentar formatos estranhos\n" + warnings
                )
            )
        return results

    # ── Camada 3: Claude ──────────────────────────────────────────────────────
    if client is None:
        return [failure(reason=f"Pandas e heurística falharam na sheet {sheet_name}; Claude não está disponível (api_key ausente)", solution="Verifique a chave de api", file_path=wb_path),]


    t0 = time.perf_counter()
    try:
        if not raw_rows:
            raw_rows = _read_raw(wb_path, sheet_name)
        tables_c = _try_claude(client, sheet_name, raw_rows)
    except Exception as e:
        return [failure(reason=f"Claude falhou na sheet {sheet_name}: {e}", solution="", file_path=wb_path),]


    for i, t in enumerate(tables_c):
        t_name = f"tabela_{i + 1}" if len(tables_c) > 1 else "tabela"
        out = _write_csv(
            output_dir, ano, mes, sheet_slug, t_name,
            t["header"], t["rows"],
        )

        results.append(
            success(
                file_path=out,
                find_cols=columns,
                execution_time_seconds=round(time.perf_counter() - t0, 4),
                warning="Foi utilizado o método heurístico, que pode apresentar formatos estranhos\n" + warnings
            )
        )

    return results


# ── Entrada pública ───────────────────────────────────────────────────────────

def convert_xls_to_csv(
    xls_path: str | Path,
    output_dir: str | Path = ".",
) -> list[Report]:
    """
    Converte um arquivo .xlsb ou .xlsx para CSVs.

    Args:
        xls_path:   Caminho para o arquivo de entrada.
        output_dir: Diretório raiz de saída (criado se não existir).
        api_key:    Chave Anthropic. Se None, lê de ANTHROPIC_API_KEY.
                    Se ausente, Claude não é acionado — erros são explícitos.

    Returns:
        Lista de TableResult (sucesso) e SheetError (falha).
        Nunca retorna silenciosamente — SheetError em vez de None.

    Raises:
        FileNotFoundError: Arquivo não encontrado.
        ValueError:        Formato de arquivo não suportado.
    """
    xls_path = Path(xls_path)
    if not xls_path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {xls_path}")

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    resolved_key = API_KEY
    client = anthropic.Anthropic(api_key=resolved_key) if resolved_key else None
    if not client:
        return   [
            failure(reason="Client do Claude Não foi definido", solution="Insira um client válido para o Claude", file_path=xls_path),
        ]
    # Data de referência: uma única chamada para o arquivo inteiro
    

    sheets = _sheet_names(xls_path)
    if not sheets:
        return   [
            failure(reason="Nenhuma aba encontrada", solution="Verifique se o arquivo está corrompido", file_path=xls_path),
        ]


    all_results: list[Report] = []
    for sheet_name in sheets:
        ano, mes = _extract_ref_date(str(xls_path), f"\n sheetname: {str(sheet_name)}", client)
        results = _process_sheet(xls_path, sheet_name, output_dir, ano, mes, client)
        all_results.extend(results)

    return all_results


# ── CLI ───────────────────────────────────────────────────────────────────────

def _main() -> None:
    reports = convert_xls_to_csv("Arquivos/V/9. 13868.7445612023-61/GM ZFM ALC - ANEXO II_2021_1Tri-Percentual de transferencia.xlsb", "",)
    with open("Relatório das Conversões PLAN.json", "w", encoding="utf-8") as arq:
        json.dump(reports, arq, indent=4, ensure_ascii=False)
    


if __name__ == "__main__":
    _main()