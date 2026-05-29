"""
xlsb_to_csv.py
==============
Lê um arquivo .xlsb, envia cada aba à API do Claude para análise
de estrutura e exporta cada tabela lógica para um arquivo CSV.

Estratégia por aba:
  • Abas pequenas (≤ SMALL_SHEET_ROWS): todas as linhas são enviadas ao
    Claude, que extrai e devolve as tabelas como JSON diretamente.
  • Abas grandes (> SMALL_SHEET_ROWS): apenas um preview é enviado; Claude
    identifica qual linha contém o cabeçalho e o índice de início dos dados;
    a extração completa é feita localmente em Python.

Dependências:
    pip install anthropic pyxlsb

Uso como módulo:
    from xlsb_to_csv import xlsb_to_csvs
    csv_paths = xlsb_to_csvs("relatorio.xlsb", output_dir="saida/")

Uso como CLI:
    python xlsb_to_csv.py relatorio.xlsb --output-dir saida/ [--api-key sk-ant-...]
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import os
import re
import sys
import time
from pathlib import Path
from typing import Any
import traceback
import anthropic
import pyxlsb
import openpyxl
from reporter_base import success, failure, Report
from utils.constants import API_KEY
# ── Configuração ──────────────────────────────────────────────────────────────

SMALL_SHEET_ROWS = 300          # abas com até N linhas → Claude extrai tudo
PREVIEW_ROWS     = 60           # linhas de preview para abas grandes
MODEL            = "claude-haiku-4-5-20251001"
MAX_TOKENS       = 8192
LOG_LEVEL        = logging.INFO



# ── Helpers internos ──────────────────────────────────────────────────────────




def _read_sheet(wb_path: str, sheet_name: str) -> list[list[Any]]:
    """Lê uma aba inteira do .xlsb ou .xlsx e retorna lista de listas (valores brutos)."""
    rows: list[list[Any]] = []
    
    # Extrai a extensão do arquivo em letras minúsculas
    _, ext = os.path.splitext(wb_path.lower())
    
    if ext == '.xlsb':
        # Lógica original para arquivos binários (.xlsb)
        with pyxlsb.open_workbook(wb_path) as wb:
            with wb.get_sheet(sheet_name) as ws:
                for row in ws.rows():
                    vals = [c.v for c in row]
                    rows.append(vals)
                    
    elif ext == '.xlsx':
        # Nova lógica para arquivos XML padrão (.xlsx)
        # read_only=True e data_only=True garantem velocidade e que fórmulas tragam apenas o valor
        wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                # row já vem como uma tupla de valores brutos graças ao values_only=True
                rows.append(list(row))
        wb.close()
        
    else:
        raise ValueError(f"Extensão de arquivo não suportada: {ext}. Use apenas .xlsb ou .xlsx")
        
    return rows


def _trim_row(row: list[Any]) -> list[Any]:
    """Remove None finais de uma linha."""
    last = -1
    for i, v in enumerate(row):
        if v is not None:
            last = i
    return row[: last + 1] if last >= 0 else []


def _rows_to_text(rows: list[list[Any]], max_rows: int | None = None) -> str:
    """Serializa linhas como texto JSON compacto (uma linha por registro)."""
    subset = rows[:max_rows] if max_rows else rows
    lines = []
    for i, row in enumerate(subset):
        trimmed = _trim_row(row)
        lines.append(f"row {i}: {json.dumps(trimmed, ensure_ascii=False, default=str)}")
    return "\n".join(lines)


def _call_claude(client: anthropic.Anthropic, prompt: str) -> str:
    """Faz uma chamada simples à API e retorna o texto da resposta."""
    msg = client.messages.create(
        model=MODEL,
        max_tokens=MAX_TOKENS,
        messages=[{"role": "user", "content": prompt}],
    )
    return msg.content[0].text


def _extract_json(text: str) -> Any:
    """Extrai o primeiro bloco JSON válido do texto de resposta do Claude."""
    # Tenta bloco de código markdown primeiro
    match = re.search(r"```(?:json)?\s*([\s\S]+?)```", text)
    if match:
        return json.loads(match.group(1))
    # Tenta JSON puro (começa com { ou [)
    match = re.search(r"(\{[\s\S]*\}|\[[\s\S]*\])", text)
    if match:
        return json.loads(match.group(1))
    raise ValueError(f"Nenhum JSON encontrado na resposta:\n{text[:500]}")


def _sanitize_filename(name: str) -> str:
    # Remove caracteres não-word exceto hífen; colapsa underscores múltiplos
    s = re.sub(r"[^\w\-]", "_", name)
    s = re.sub(r"_+", "_", s).strip("_")
    return s or "tabela"

def _retorna_data(path, ):
    client = anthropic.Anthropic(api_key=API_KEY)

    prompt = f""""
    A partir do nome do arquivo "{path}"
    retorne o mẽs e ano de referẽncia no formato YYYY_MM
    Não retorne comentários ou explicações, apenas o YYYY_MM
    Se não encontrar o ano ou o mês, retorne substituindo o que faltar por 00. Exemplo com mês não encontrado: 2021_00
    Se o ano for um range, retorne a data no formato YYYY-YYYY_00
    Formatos de retorno válidos:
    1. 2022_01
    2. 2003_12
    3. 2019_00
    4. 0000_06
    5. 2019-2022_00
    """

    message = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=7,
        messages=[{"role": "user", "content": prompt}],
    )
    return message.content[0].text.strip()

# ── Lógica de extração por tipo de aba ────────────────────────────────────────

def _process_small_sheet(
    client: anthropic.Anthropic,
    sheet_name: str,
    rows: list[list[Any]],
    output_dir: Path,
    input_path:Path
) -> list[Report]:
    """
    Para abas pequenas: envia todas as linhas ao Claude e pede que ele
    extraia e devolva as tabelas como JSON.

    Claude deve retornar:
    {
      "tables": [
        {
          "name": "nome_da_tabela",
          "columns": ["col1", "col2", ...],
          "rows": [[val, val, ...], ...]
        }
      ]
    }
    """
    rows_text = _rows_to_text(rows)

    prompt = f"""Você receberá o conteúdo bruto de uma aba de planilha chamada "{sheet_name}".
Cada linha está no formato: row <índice>: [valores...]

Analise a estrutura e extraia TODAS as tabelas lógicas presentes.
Ignore linhas de título/cabeçalho descritivo (nomes de empresa, períodos, notas de rodapé, etc.).
Inclua apenas dados tabulares com cabeçalhos e registros.

Retorne SOMENTE um JSON válido, sem texto adicional, no seguinte formato:
{{
  "tables": [
    {{
      "name": "<nome_descritivo_sem_espacos>",
      "columns": ["<col1>", "<col2>", ...],
      "rows": [
        [<val1>, <val2>, ...],
        ...
      ]
    }}
  ]
}}

Conteúdo da aba:
{rows_text}
"""
    response = _call_claude(client, prompt)

    try:
        spec = _extract_json(response)
    except Exception as e:
        return [
            failure(
                reason=f"Erro ao receber resposta do claude sobre planilha pequena. \n O prompt enviado foi {prompt} \n\n\n E a resposta do claude foi {response}",
                solution="",
                file_path=str(input_path)
            ),
        ]
    tables = spec.get("tables", [])

    reports:list[Report] = []

    written: list[str] = []
    sheet_slug = _sanitize_filename(sheet_name)
    for table in tables:
        start_time = time.perf_counter()
        name    = _sanitize_filename(table.get("name", "tabela"))
        columns = table.get("columns", [])
        data    = table.get("rows", [])
        if not data:
            reports.append(failure(
                    reason =f"Aba '{name}' sem dados — pulando.",
                    solution="Verifique se essa aba da planilha realmente possui dados e se não estão corrompidos",
                    file_path=Path(input_path)
                ))
            continue

        data_referencia = _retorna_data(input_path, )
        ano = data_referencia.split("_")[0]
        mes = data_referencia.split("_")[1]
        root_output_dir = output_dir / ano / mes
        os.makedirs(str(root_output_dir), exist_ok=True)

        out_path =  root_output_dir / f"{sheet_slug}__{name}.csv"
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f,delimiter=";")
            if columns:
                writer.writerow(columns)
            writer.writerows(data)
        written.append(str(out_path))
        reports.append(success(
            file_path=Path(out_path),
            find_cols=columns,
            execution_time_seconds = round(time.perf_counter() - start_time, 4),
        ))
        
    return reports



def _process_large_sheet(
    client: anthropic.Anthropic,
    sheet_name: str,
    rows: list[list[Any]],
    output_dir: Path,
    input_path:Path
) -> list[str]:
    """
    Para abas grandes: envia preview ao Claude para identificar a estrutura,
    depois faz a extração local de todas as linhas.

    Claude deve retornar:
    {
      "tables": [
        {
          "name": "nome_da_tabela",
          "header_row": <índice>,
          "data_start_row": <índice>,
          "data_end_row": <índice ou null>
        }
      ]
    }
    """
    preview_text = _rows_to_text(rows, max_rows=PREVIEW_ROWS)

    prompt = f"""Você receberá um preview do início de uma aba de planilha chamada "{sheet_name}".
Cada linha está no formato: row <índice>: [valores...]

Identifique TODAS as tabelas tabulares presentes (pode haver mais de uma).
Para cada tabela, informe:
  - header_row: índice da linha que contém os cabeçalhos das colunas
  - data_start_row: índice da primeira linha de dados (logo após o cabeçalho)
  - data_end_row: índice da última linha de dados (null se vai até o fim da aba)

Retorne SOMENTE um JSON válido, sem texto adicional:
{{
  "tables": [
    {{
      "name": "<nome_descritivo_sem_espacos>",
      "header_row": <int>,
      "data_start_row": <int>,
      "data_end_row": <int ou null>
    }}
  ]
}}

Preview da aba:
{preview_text}
"""
    response = _call_claude(client, prompt)

    spec = _extract_json(response)
    tables = spec.get("tables", [])

    written: list[str] = []
    sheet_slug = _sanitize_filename(sheet_name)

    reports:list[Report] = []

    for table in tables:
        start_time = time.perf_counter()
        name       = _sanitize_filename(table.get("name", "tabela"))
        header_row = table.get("header_row")
        data_start = table.get("data_start_row")
        data_end   = table.get("data_end_row")  # None = até o fim

        if header_row is None or data_start is None:
            reports.append(failure(
                    reason =f"Aba '{name}' sem header_row/data_start_row — pulando.",
                    solution="Verifique se essa aba da planilha realmente possui dados e se não estão corrompidos",
                    file_path=input_path
                ))
            continue

        header = [str(v) if v is not None else f"CAMPO_{i}"
                  for i, v in enumerate(_trim_row(rows[header_row]))]


        data_referencia = _retorna_data(input_path,)
        ano = data_referencia.split("_")[0]
        mes = data_referencia.split("_")[1]

        root_output_dir = output_dir / ano / mes
        os.makedirs(str(root_output_dir), exist_ok=True)

        out_path =  root_output_dir / f"{sheet_slug}__{name}.csv"
        data_rows_written = 0
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow(header)
            for i in range(data_start, len(rows)):
                if data_end is not None and i > data_end:
                    break
                trimmed = rows[i][: len(header)]
                # Pula linhas completamente vazias
                if any(v is not None for v in trimmed):
                    writer.writerow(trimmed)
                    data_rows_written += 1

        written.append(str(out_path))
        reports.append(success(
            file_path=Path(out_path),
            find_cols=header,
            execution_time_seconds = round(time.perf_counter() - start_time, 4),
        ))

    return reports


# ── Função principal ──────────────────────────────────────────────────────────

def xls_to_csvs(
    xls_path: str,
    output_dir: str = ".",
    verbose: bool = False,
) -> list[str]:
    """
    Lê um arquivo .xlsb, usa a API do Claude para identificar a estrutura
    de cada aba e exporta cada tabela lógica para um arquivo CSV.

    Args:
        xls_path:  Caminho para o arquivo .xlsb.
        output_dir: Diretório onde os CSVs serão salvos (criado se não existir).
        api_key:    Chave da API Anthropic. Se None, usa ANTHROPIC_API_KEY.
        verbose:    Ativa logging em nível DEBUG.

    Returns:
        Lista com os caminhos absolutos dos CSVs gerados.

    Raises:
        FileNotFoundError: Se o arquivo .xlsb não existir.
        anthropic.AuthenticationError: Se a chave da API for inválida.
        ValueError: Se Claude retornar uma resposta sem JSON válido.
    """
    xls_path = Path(xls_path)
    if not xls_path.exists():
        return [
            failure(
                reason=f"Arquivo não encontrado: {xls_path}",
                solution="Verifique se o arquivo realmente existe.",
                file_path=xls_path
            ),
        ]

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)


    client = anthropic.Anthropic(api_key=API_KEY)

    all_csv_paths: list[str] = []


    ext = xls_path.suffix.lower()
    if ext == ".xlsb":
        with pyxlsb.open_workbook(str(xls_path)) as wb:
            sheet_names = wb.sheets
    elif ext == ".xlsx":
        wb = openpyxl.load_workbook(str(xls_path), read_only=True)
        sheet_names = wb.sheetnames
    else:
        return [
            failure(
                reason=f"Arquivo não possui um formato válido (xlsx, xlsb): {xls_path}",
                solution="Verifique se o arquivo realmente existe e se o formato está correto.",
                file_path=xls_path
            ),
        ]

    try:

        all_reports:list[Report] = []

        for sheet_name in sheet_names:
            rows = _read_sheet(str(xls_path), sheet_name)

            if len(rows) <= SMALL_SHEET_ROWS:
                sheet_report = _process_small_sheet(client, sheet_name, rows, output_dir, input_path=Path(xls_path))
            else:
                sheet_report = _process_large_sheet(client, sheet_name, rows, output_dir, input_path=Path(xls_path))
            
            all_reports += sheet_report


        if len(all_reports) < 1:
            return [
                failure(
                    reason=f"Não foram encontradas abas no arquivo.",
                    solution="Verifique se o arquivo realmente existe e se o formato está correto.",
                    file_path=xls_path
                ),
            ]
        return all_reports
    except Exception as e:
        return [
            failure(
                reason=f"Erro desconhecido ao tratar o arquivo: {traceback.format_exc()}",
                solution="Verifique se o arquivo realmente existe e se o formato está correto.",
                file_path=xls_path
            ),
        ]




def parse_xls_to_csv(xls_path, output_dir) -> list[dict]:
    try:
        paths = xls_to_csvs(xls_path=xls_path,output_dir=output_dir,)
    except FileNotFoundError as e:
        return [
            failure(
                reason="File Not Found", 
                solution="Verifique se o arquivo existe e está no formato correto."
            )
        ]
    except Exception as e:
        return [
            failure(
                reason="Falha desconhecida: " + str(e), 
                solution="Verifique se o arquivo existe e está no formato correto."
            )
        ]

if __name__ == "__main__":
    reports = xls_to_csvs(xls_path="Arquivos/III/Customs - Exportação/Export 2020.xlsx",output_dir="Tratados",)
